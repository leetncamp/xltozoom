#!/usr/bin/env python

import pytz
from pdb import set_trace as debug
from argparse import ArgumentParser
from openpyxl import load_workbook
from create_or_update_zoom import create_or_update_zoom
import traceback
import os
import sys
import psutil


parser = ArgumentParser()
parser.add_argument("--clearAll", action="store_true", help="delete all meetings listed in schedule.xlsx",)
parser.add_argument("--users", action="store_true")
parser.add_argument("--file", default="schedule.xlsx")
ns = parser.parse_args()

UTC = pytz.timezone("UTC")

my_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(my_dir, "zoomus"))

excelpath = os.path.join(my_dir, ns.file)

wb = load_workbook(excelpath)
ws = wb.active

dirty = False

headers = [item.value for item in ws[1]]


def excel_is_open():
    excel = [item.name() for item in psutil.process_iter() if "excel" in item.name().lower()]
    return (bool(excel))


def save_excel(msg):

    if excel_is_open():
        from tkinter import messagebox, Tk
        import tkinter
        TK_SILENCE_DEPRECATION = 1
        window = tkinter.Tk()
        window.wm_withdraw()
        messagebox.showinfo("Warming", msg)
    wb.save(excelpath)


for row in ws.iter_rows(min_row=2):
    try:
        data = dict(zip(headers, [item.value for item in row]))
        integration = data.get("integration")
        if integration in ["Zoom", None]:
            result = create_or_update_zoom(data)
            print("{0}: {1}".format(result.get("action").capitalize(), result.get("topic")))
            action = result.get("action")
            #if data.get("peerreviewid") == "Workshopapplication-36":
            #    debug()

            if action == "create":
                zoomid = row[headers.index("zoomid")]
                zoomid.value = str(result.get("id"))
                print("Created {0}".format(result.get("topic")))
                if not dirty:
                    dirty = True
            elif action == "skipped":
                print(action)
    except:
        tb = traceback.format_exc()
        if dirty:
            save_excel("Warning, An exception occurred. However, before that exception occurred, the file, {},  was modifed by this script with updated information about zoom meetings. Please close the file in Excel without saving if it is open. \n\n{}".format(excelpath, tb))
        else:
            save_excel("Warning, An exception occurred.  \n\n{}".format(tb))


if dirty:
    save_excel("{} has been updated. Please reopen it without saving.".format(excelpath))
