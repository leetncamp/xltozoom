#!/usr/bin/env python
 
import pytz
from pdb import set_trace as debug
from argparse import ArgumentParser
from openpyxl import load_workbook
from create_or_update_zoom import create_or_update_zoom
import traceback
import os
import sys
import psutil
from zsecrets import graphical_warning

parser = ArgumentParser()
parser.add_argument("--clearAll", action="store_true", help="delete all meetings listed in schedule.xlsx",)
parser.add_argument("--users", action="store_true")
parser.add_argument("--file", default="schedule.xlsx")
ns = parser.parse_args()

UTC = pytz.timezone("UTC")

my_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(my_dir, "zoomus"))

excelpath = os.path.join(my_dir, ns.file)

wb = load_workbook(excelpath)
ws = wb.active

dirty = False

headers = [item.value for item in ws[1]]


def excel_is_open():
    return True
    excel = [item.name() for item in psutil.process_iter() if "excel" in item.name().lower()]
    return bool(excel)


def save_excel(msg):
    if excel_is_open():
        if graphical_warning:
            from tkinter import messagebox
            import tkinter
            window = tkinter.Tk()
            window.wm_withdraw()
            messagebox.showinfo("Warming", msg)
        else:
            print(msg)
            input("Press Enter to save schedule.xlsx with updated information.  Press ^c otherwise. ")
    wb.save(excelpath)
    sys.exit()


for row in ws.iter_rows(min_row=2):
    try:
        data = dict(zip(headers, [item.value for item in row]))
        integration = data.get("integration")
        if integration in ["Zoom", None]:
            zoom_username = data.get("zoom_username")
            print(zoom_username)
            if not zoom_username:
                print("Skipping row {} because zoom_username is empty".format(data))
            result = create_or_update_zoom(data)

            action = result.get("action")
            if action == "skip":
                continue
            zoom_id = row[headers.index("zoom_id")]
            zoom_id.value = str(result.get("zoom_id"))

            zoom_join_link = result.get("join_url")

            """Update the join link each time because the link can change if you change the attributes of the meeting 
            """
            
            if zoom_join_link:
                join_link_cell = row[headers.index("zoom_join_link")]
                join_link_cell.value = zoom_join_link

            zoom_start_link = result.get("start_url")


            """Update the join link each time because the link can change if you change the attributes of the meeting"""
            
            if zoom_start_link:
                start_link_cell = row[headers.index("zoom_start_link")]
                start_link_cell.value = zoom_start_link

            if not dirty:
                dirty = True
            
    except:
        tb = traceback.format_exc()
        if dirty:
            save_excel("Warning, An exception occurred. However, before that exception occurred, "
                       "the file, {},  was modified by this script with updated information about zoom meetings. "
                       "Please close the file in Excel without saving if it is open. \n\n{}".format(excelpath, tb))
        else:
            save_excel("Warning, An exception occurred.  \n\n{}".format(tb))


if dirty:
    save_excel("{} has been updated. Please reopen it without saving.".format(excelpath))
