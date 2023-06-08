#!/usr/bin/env python
# encoding: utf-8

"""

This script requires MacOS and Apple Mail to work without modification.  When creating one Zoom account per poster, you
will need to create 2000 accounts (NeurIPS) and manually activate them.   This script helps automate that with selenium
driver.

pip install selenium

There are more installation steps required as documented here:
https://selenium-python.readthedocs.io/getting-started.html
On MacOs, the driver will not be signed and you'll need to  xattr -d com.apple.quarantine chromedriver

This script is set up to use the Gecko Driver and Firefox, but can be easily modified to use another browser.  Safari on
MacOS includes a selenium driver and works without much fuss.  To use this script:

0. Set the meeting name below
1. Select, then drag and drop all zoom activation emails from Apple Mail to a folder.  This will produce one .eml file
   per zoom activation email.

2. Pass that folder location as a parameter to this script. An optional 2nd parameter is the schedule.xlsx file from
   xlttozoom. If this file contains zoom_host_password  zoom_username and type fields, then these will be used
   rather than the hard coded defaults for the account password. The event type will be used as the last name.

The script will open each email, find the activation link url, open it in a browser, and set the password to the value
defined in teh zoompwd variable below.

Example:

./activate_zoom_accounts.py ~/Desktop/zoom_activation_emls_folder xltozoom/schedule.xlsx

"""
import time

import selenium
from selenium.webdriver.common.by import By

meeting_name = "CVPR 2023"
input(f"Using {meeting_name}: ")

import os
from pdb import set_trace as debug
stop = debug
import glob
from argparse import ArgumentParser
from email.parser import BytesParser
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from msedge.selenium_tools import Edge, EdgeOptions

zoompwd = os.getenv("ZOOMPWD")

parser = ArgumentParser()
parser.add_argument("file", nargs="*", help="supply path to folder to collection of account activation .eml (emails dragged from Apple mail) files from zoom as first param. The 2nd param should be the schedule xlsx that has zoom_host_password, host_zoom_user_email and type.")
ns = parser.parse_args()

emls = glob.glob(ns.file[0] + "/*.eml")
email_parser = BytesParser


def viewhtml(html):
    open("/tmp/delme.html", 'wb').write(html)
    os.system("open /tmp/delme.html")


data = {}

if len(ns.file) > 1:
    schedulexlsx = ns.file[1]
    wb = load_workbook(schedulexlsx)
    ws = wb.active
    headers = [i.value for i in ws[1] if i.value]

    for row in ws.iter_rows(min_row=2):
        row_data = dict(zip(headers, [i.value for i in row]))
        zoom_username = row_data.get("zoom_username")
        row_type = row_data.get("type")
        zoom_host_password = row_data.get("zoom_host_password")
    
        if zoom_host_password and row_type and zoom_username:
            data[zoom_username] = {
                "zoom_host_password": zoom_host_password,
                "type": row_type,
                }
        else:
            print("skipping row {}".format(row_data))

default_password = "CubeVis31"
default_type = "Talk"

schedule_data = data

for eml in emls:

    fp = open(eml, 'rb')
    msg = BytesParser().parse(fp)
    html = msg.get_payload(decode=True).decode("UTF-8")
    soup = bs(html, features="lxml")
    links = soup.findAll("a")
    login_email = soup.find("table").find("table").find("table").findAll("td")[0].text.strip().replace("Hello ", "").strip(",")
    print(login_email)
    schedulexlsx_info = schedule_data.get(login_email, {})

    Password = schedulexlsx_info.get("zoom_host_password", default_password)
    Lastname = schedulexlsx_info.get("type", default_type)

    # Exclude links that aren't activation links
    links = [link['href'] for link in links if "zoom.us/activate" in link['href']]
    # There may be more than one activation link; get only the first one.

    if links:
        link = links[0]
        driver = webdriver.Firefox()
        #driver = webdriver.Chrome()
        #options = EdgeOptions()
        #options.use_chromium = True
        #driver = Edge(options=options)
        driver.get(link)
        #try:
        #    login = driver.find_element_by_link_text("Sign Up with a Password")
        #except NoSuchElementException:
        #    print("{} has already been processed. Press c [enter] to continue".format(login_email))
        #    os.rename(eml, eml.replace(".eml", ".txt"))
#
        #    driver.close()
        #    continue
        #print("Activating {} with {} and lastname of {}".format(login_email, Password, Lastname))
        #login.click()
        # Try to get the first name. If it's not there it's usually because the link has expired or the
        # profile has already been activated.
        try:
            firstname = driver.find_element_by_id("firstName")
        except selenium.common.exceptions.NoSuchElementException:
            try:
                link_expired = driver.find_element(By.CLASS_NAME, 'error-message').text
                print(f"{link_expired}")
                driver.close()
                continue
            except selenium.common.exceptions.NoSuchElementException:
                try:
                    already_activated = driver.find_element(By.CLASS_NAME, 'zm-signup-layout__desc').text
                    print(f"{already_activated}")
                    driver.close()
                    continue
                except selenium.common.exceptions.NoSuchElementException:
                    print(f"An unknown error occurred with processing {eml}")
                    debug()
                    continue
        firstname.send_keys(meeting_name)
        lastname = driver.find_element_by_id("lastName")
        lastname.send_keys(Lastname)
        password = driver.find_element_by_name("password")
        password.send_keys(Password)
        cpassword = driver.find_element_by_name("confirmPassword")
        cpassword.send_keys(Password)
        cpassword.send_keys("\t")
        time.sleep(1)
        Continue = driver.find_element_by_xpath("//span[contains(text(), 'Continue')]")
        Continue.click()
        driver.close()
        print(f"Activated {login_email}")
        continue
